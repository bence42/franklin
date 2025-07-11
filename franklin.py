import pandas as pd
import argparse
import os
import glob
from typing import List
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class Sheet():
    def __init__(self, name: str, data: pd.DataFrame):
        self.name = name
        self.data = data

    def filter(self, condition) -> 'Sheet':
        """ e.g. sheet.filter(lambda df: df["Variant_Frequency"] >= 0.35) """
        self.data = self.data[condition(self.data)]
        return self

    def sort(self, *args, **kwargs):
        """ Wrapper around pd.DataFrame's sort_values. """
        self.data.sort_values(*args, **kwargs,  inplace=True)


def autofit_columns(ws: Worksheet):
    """ Sets each columns width so that the longest text still fits. """
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add padding
        ws.column_dimensions[column].width = adjusted_width

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def order_columns(df: pd.DataFrame) -> pd.DataFrame:
    """ Reorder columns in the specified order."""
    new_order = [
        'Chromosome',
        'Position',
        'Reference_Allele',
        'Variant_Allele',
        'Clinvar_Significance',
        'Gene_ID',
        'Clinvar_ID',
        'Variant_Frequency',
        'Total_Depth',
        'Variant_Type',
        'Consequence',
        'dbSNP_ID',
        'Hgvsg_ID',
        'OMIM_Link',
        'GenotypeQuality',
        'Genotype',
        'Filters',
        'Ref_Allele_Depth',
        'Variant_Allele_Depth',
        'Strand_Bias',
        'All_Freq_All',
        'Non-Finnish_Eur_Allele_Freq',
        'EastAsian_Allele_Freq',
        'SouthAsian_Allele_Freq',
        'Latino_Allele_Freq',
        'African_Allele_Freq'
    ]

    df = df[new_order]
    return df


def insert_franklin_column(df: pd.DataFrame):
    """" Inserts a new column named Frankin right to Clinvar_ID column. """
    col_idx = df.columns.get_loc('Clinvar_ID') + 1
    df.insert(col_idx, 'Franklin', value=None)


def highlight_Clinvar_Significance(wb: Workbook, sheet_name: str, clinvar_idx: int, franklin_idx: int):
    """ Colors cells based on Clinvar_Significance on a single sheet. """

    text_to_highlight = 'benign'
    max_row_in_sheet = wb[sheet_name].max_row
    clinvar_col = get_column_letter(col_idx=clinvar_idx)

    def highlight_benign():
        """ Colors Clinvar_Significance green when containing text 'benign'. """
        green_fill = DifferentialStyle(fill=PatternFill(bgColor="92d050"))
        benign_rule = Rule(type="containsText",
                        operator="containsText",
                        text=text_to_highlight,
                        dxf=green_fill)


        clinvar_col_range = f'{clinvar_col}1:{clinvar_col}{max_row_in_sheet}'

        wb[sheet_name].conditional_formatting.add(
            range_string=clinvar_col_range, cfRule=benign_rule)

    def highlight_not_benign():
        """ Colors 'Franklin' cells yellow when the matching Clinvar_Significance does not contain 'benign'. """
        yellow_fill = DifferentialStyle(fill=PatternFill(bgColor="ffff00"))
        not_benign_rule = Rule(type="expression",
                            formula=[
                                # not empty and does not contain 'benign'. applied from row 2
                                f'ISERROR(SEARCH("{text_to_highlight}", ${clinvar_col}2))'],
                            dxf=yellow_fill)

        franklin_col = get_column_letter(col_idx=franklin_idx)
        franklin_col_range = f'{franklin_col}2:{franklin_col}{max_row_in_sheet}'

        wb[sheet_name].conditional_formatting.add(
            range_string=franklin_col_range, cfRule=not_benign_rule)
        
    highlight_benign()
    highlight_not_benign()


def process_file(input_file: str):
    variants = Sheet('variants', pd.read_csv(input_file, sep="\t"))
    variants.sort('Variant_Frequency', ascending=False)

    variants.data = order_columns(variants.data)
    insert_franklin_column(variants.data)

    extended = Sheet('extended', variants.data.copy())
    extended.filter(lambda df: df['Variant_Frequency'] >= 0.35)
    extended.sort('Clinvar_Significance')

    clinical = Sheet('klinikai', variants.data.copy())
    clinical.sort(['Gene_ID', 'Variant_Frequency', 'Clinvar_Significance'])
    interesting_genes = ['BRCA1', 'BRCA2', 'PALB2', 'ATM', 'MLH1', 'MSH2', 'MSH6', 'PMS2', 'EPCAM', 'stk11']
    clinical.filter(lambda df: df['Gene_ID'].isin(interesting_genes))

    input_file_name = os.path.splitext(input_file)[0]
    excel_file = f'{input_file_name}.xlsx'
    wb = save_to_xlsx(filename=excel_file, 
                      sheets=[variants,
                              extended,
                              clinical])

    highlight_Clinvar_Significance(
        wb=wb,
        sheet_name=extended.name,
        clinvar_idx=extended.data.columns.get_loc('Clinvar_Significance') + 1,
        franklin_idx=extended.data.columns.get_loc('Franklin') + 1)

    highlight_Clinvar_Significance(
        wb=wb,
        sheet_name=clinical.name,
        clinvar_idx=clinical.data.columns.get_loc('Clinvar_Significance') + 1,
        franklin_idx=clinical.data.columns.get_loc('Franklin') + 1)

    for ws in wb.worksheets:
        autofit_columns(ws)
        ws.auto_filter.ref = ws.dimensions

    wb.save(excel_file)


def save_to_xlsx(filename: str, sheets: List[Sheet]) -> Workbook:
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    for sheet in sheets:
        sheet.data.to_excel(writer, index=False, sheet_name=sheet.name)
    writer.close()

    return load_workbook(filename)


def process_files(paths: List[str]):
    files = []
    for path in paths:
        if os.path.isdir(path):
            files.extend(glob.glob(f'{path}/*.txt'))
        else:
            files.append(path)

    for file in files:
        print(f'processing {file}')
        if not os.path.exists(file):
            print(f"> error: '{file}' does not exist")
            exit(1)
        process_file(file)


def main():
    parser = argparse.ArgumentParser(description='Tool')

    parser.add_argument('-i',
                        '--inputs',
                        nargs='+',
                        required=True,
                        help='input file(s) or folder(s), space spearated')

    args = parser.parse_args()

    process_files(args.inputs)


main()
